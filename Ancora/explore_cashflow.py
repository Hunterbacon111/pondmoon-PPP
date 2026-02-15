"""
Explore the Ancora Cash Flow Projections Excel file.
Lists all sheets, prints first 60 rows of each, shows headers and data types.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys

FILE_PATH = "/Users/fangheli/Downloads/Property_Analysis_Tool_ClaudeCode/Ancora/Data_Annual_Budget/Ancora_Cash Flow Projections_1.23.2026.xlsx"

wb = openpyxl.load_workbook(FILE_PATH, data_only=True)

print("=" * 120)
print(f"FILE: {FILE_PATH}")
print(f"TOTAL SHEETS: {len(wb.sheetnames)}")
print(f"SHEET NAMES: {wb.sheetnames}")
print("=" * 120)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'#' * 120}")
    print(f"SHEET: '{sheet_name}'")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    print(f"{'#' * 120}")

    # Print merged cells info
    if ws.merged_cells.ranges:
        print(f"\nMerged cell ranges ({len(ws.merged_cells.ranges)}):")
        for mc in list(ws.merged_cells.ranges)[:30]:
            print(f"  {mc}")
        if len(ws.merged_cells.ranges) > 30:
            print(f"  ... and {len(ws.merged_cells.ranges) - 30} more")

    # Print first 60 rows with all columns
    print(f"\n--- First 60 rows (all {ws.max_column} columns) ---")

    max_rows_to_print = min(60, ws.max_row)

    # Build column header line
    col_headers = ["Row#"]
    for c in range(1, ws.max_column + 1):
        col_headers.append(f"{get_column_letter(c)}({c})")
    print(" | ".join(col_headers))
    print("-" * 120)

    for row_idx in range(1, max_rows_to_print + 1):
        row_values = [f"R{row_idx:3d}"]
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None:
                row_values.append("")
            else:
                # Truncate long strings
                s = str(val)
                if len(s) > 40:
                    s = s[:37] + "..."
                row_values.append(s)
        print(" | ".join(row_values))

    # Data type analysis for each column
    print(f"\n--- Column Data Type Analysis (sampling first 60 rows) ---")
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        types_found = {}
        sample_values = []
        for row_idx in range(1, max_rows_to_print + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            t = type(val).__name__
            types_found[t] = types_found.get(t, 0) + 1
            if val is not None and len(sample_values) < 5:
                s = str(val)
                if len(s) > 50:
                    s = s[:47] + "..."
                sample_values.append(s)
        type_str = ", ".join(f"{k}:{v}" for k, v in types_found.items())
        print(f"  Col {col_letter}({col_idx}): types=[{type_str}]  samples={sample_values}")

    # If the sheet has more than 60 rows, also show rows 61-120 or up to max
    if ws.max_row > 60:
        extra_end = min(120, ws.max_row)
        print(f"\n--- Rows 61 to {extra_end} ---")
        for row_idx in range(61, extra_end + 1):
            row_values = [f"R{row_idx:3d}"]
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    row_values.append("")
                else:
                    s = str(val)
                    if len(s) > 40:
                        s = s[:37] + "..."
                    row_values.append(s)
            print(" | ".join(row_values))

    # If even more rows, show the last 20 rows
    if ws.max_row > 120:
        print(f"\n--- Last 20 rows (rows {ws.max_row - 19} to {ws.max_row}) ---")
        for row_idx in range(max(121, ws.max_row - 19), ws.max_row + 1):
            row_values = [f"R{row_idx:3d}"]
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    row_values.append("")
                else:
                    s = str(val)
                    if len(s) > 40:
                        s = s[:37] + "..."
                    row_values.append(s)
            print(" | ".join(row_values))

    print(f"\n{'=' * 120}")

wb.close()
print("\nDone.")
