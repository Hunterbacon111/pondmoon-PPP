"""Extract budget data from the Ancora Cash Flow Projections XLSX."""
import os
import glob
from openpyxl import load_workbook
from src.config import (DATA_BUDGET, BUDGET_MONTHS, BUDGET_ROW_MAP,
                         BUDGET_DATA_COL_START, BUDGET_TOTAL_COL)


def _find_budget_file():
    """Find the most recent budget XLSX file."""
    pattern = os.path.join(DATA_BUDGET, "**", "*.xlsx")
    files = glob.glob(pattern, recursive=True)
    files = [f for f in files if not os.path.basename(f).startswith("~")]
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def _parse_value(raw):
    """Parse a cell value into a float."""
    if raw is None:
        return None
    try:
        val = float(raw)
        return val
    except (ValueError, TypeError):
        return None


def extract():
    """Extract monthly budget data from the Cash Flow Projections XLSX.

    Returns a dict with year, months list, and metrics dict.
    Each metric key maps to a list of 12 monthly values.
    """
    filepath = _find_budget_file()
    if not filepath:
        print("  [budget] No budget file found.")
        return {"year": None, "months": BUDGET_MONTHS, "metrics": {}, "source_file": None}

    print(f"  [budget] Reading: {os.path.basename(filepath)}")

    wb = load_workbook(filepath, data_only=True)
    # Use the first sheet (Ext_Capital_Call)
    ws = wb.active
    print(f"  [budget] Sheet: {ws.title}")

    metrics = {}
    for metric_key, row_num in BUDGET_ROW_MAP.items():
        monthly_vals = []
        for col_offset in range(12):
            col_idx = BUDGET_DATA_COL_START + col_offset  # F=6 through Q=17
            val = ws.cell(row=row_num, column=col_idx).value
            monthly_vals.append(_parse_value(val))
        metrics[metric_key] = monthly_vals

        # Also grab the annual total from column R
        total_val = ws.cell(row=row_num, column=BUDGET_TOTAL_COL).value
        metrics[f"{metric_key}_annual"] = _parse_value(total_val)

    wb.close()

    # Compute annual totals for key metrics
    annual_totals = {}
    for key in ["total_income", "total_opex", "noi", "net_income", "debt_service",
                "payroll_benefits", "marketing", "contract_services", "insurance",
                "management_fees", "taxes", "ground_lease", "utilities",
                "total_net_rental_income", "vacancy_loss", "concessions"]:
        # Prefer the _annual value from the spreadsheet total column
        annual_key = f"{key}_annual"
        if annual_key in metrics and metrics[annual_key] is not None:
            annual_totals[key] = metrics[annual_key]
        else:
            vals = metrics.get(key, [])
            non_null = [v for v in vals if v is not None]
            annual_totals[key] = sum(non_null) if non_null else None

    return {
        "year": 2026,
        "months": BUDGET_MONTHS,
        "metrics": metrics,
        "annual_totals": annual_totals,
        "source_file": os.path.basename(filepath),
    }
