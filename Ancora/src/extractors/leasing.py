"""Extract weekly leasing data from Ancora Weekly Leasing Update XLSX."""
import os
import glob
from datetime import datetime
from openpyxl import load_workbook
from src.config import DATA_LEASING


def _parse_xlsx_weekly(filepath):
    """Parse the Ancora weekly leasing tracking XLSX file.

    Structure: sheet "Ancora", 220 total units
    - Row 3 = dates (columns C onward, newest first)
    - Row 4 = Leased #
    - Row 5 = Leased %
    - Row 6 = Occupied #
    - Row 7 = Occupied %
    - Row 8 = Move Out
    - Row 9 = Move In
    - Row 10 = New Prospects
    - Row 11 = Walk In traffic
    - Row 12 = Gross Leases
    - Row 13 = Net Leases
    - Row 14 = 30-day occp trend
    - Row 15 = 60-day occp trend
    - Row 16 = PSF - All Leases
    - Row 17 = All Executed Effective
    - Rows 20-21 = Concessions
    - Rows 25-26 = Construction notes
    - Rows 31-33 = Delinquency
    """
    wb = load_workbook(filepath, data_only=True)

    # Find the right sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "ancora" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    print(f"  [leasing] Using sheet: {sheet_name}")

    # --- Top table: rows 3-17, columns C onward ---
    dates = []
    col_start = 3  # column C
    for col in range(col_start, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val and isinstance(val, datetime):
            dates.append((col, val))

    weekly_metrics = []
    for col, dt in dates:
        def cell_val(row, c=col):
            return ws.cell(row=row, column=c).value

        entry = {
            "week_ending": dt.strftime("%Y-%m-%d"),
            "leased_num": cell_val(4),
            "leased_pct": cell_val(5),
            "occupied_num": cell_val(6),
            "occupied_pct": cell_val(7),
            "move_out": cell_val(8),
            "move_in": cell_val(9),
            "new_prospects": cell_val(10),
            "walk_in_traffic": cell_val(11),
            "gross_leases": cell_val(12),
            "net_leases": cell_val(13),
            "trend_30_day": cell_val(14),
            "trend_60_day": cell_val(15),
            "psf_all_leases": cell_val(16),
            "psf_executed_effective": cell_val(17),
        }

        # Convert percentages stored as decimals (e.g. 0.0955 -> 9.55)
        for key in ["leased_pct", "occupied_pct", "trend_30_day", "trend_60_day"]:
            if entry[key] is not None and isinstance(entry[key], (int, float)):
                if entry[key] < 1:  # stored as decimal
                    entry[key] = round(entry[key] * 100, 2)
                else:
                    entry[key] = round(entry[key], 2)

        # Round unit counts
        for key in ["leased_num", "occupied_num"]:
            if entry[key] is not None and isinstance(entry[key], (int, float)):
                entry[key] = round(entry[key])

        weekly_metrics.append(entry)

    # Sort chronologically (oldest first)
    weekly_metrics.sort(key=lambda w: w["week_ending"])

    # --- Concessions (rows 20-21) ---
    concessions = []
    for row in range(20, 22):
        val = ws.cell(row=row, column=2).value  # column B
        if val and isinstance(val, str) and val.strip():
            concessions.append(val.strip())

    # --- Construction notes (rows 25-26) ---
    construction_notes = []
    for row in range(25, 27):
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, str) and val.strip():
            construction_notes.append(val.strip())

    # --- Delinquency (rows 31-33) ---
    delinquency_notes = []
    for row in range(31, 34):
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, str) and val.strip():
            label = val.strip()
            if label.lower() in ("delinquency", "delinquency:"):
                continue
            delinquency_notes.append(label)

    wb.close()

    return {
        "weekly_metrics": weekly_metrics,
        "concessions": concessions,
        "construction_notes": construction_notes,
        "delinquency_notes": delinquency_notes,
    }


def extract():
    """Scan Data_Leasing directory and extract weekly data.

    Primary source: Weekly Leasing Update XLSX.
    """
    if not os.path.exists(DATA_LEASING):
        print("  [leasing] Data_Leasing directory not found.")
        return {"weeks": [], "xlsx_data": None}

    # Find XLSX files
    xlsx_files = glob.glob(os.path.join(DATA_LEASING, "**", "*.xlsx"), recursive=True)
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith("~")]

    xlsx_data = None
    if xlsx_files:
        xlsx_path = max(xlsx_files, key=os.path.getmtime)
        print(f"  [leasing] Parsing XLSX: {os.path.basename(xlsx_path)}")
        xlsx_data = _parse_xlsx_weekly(xlsx_path)
        print(f"  [leasing] Found {len(xlsx_data['weekly_metrics'])} weekly snapshots")
    else:
        print("  [leasing] No XLSX file found.")
        return {"weeks": [], "xlsx_data": None}

    # Build unified output
    weeks = []
    for wm in xlsx_data["weekly_metrics"]:
        week_entry = {
            "week_ending": wm["week_ending"],
            "source_type": "xlsx",
            "occupancy_pct": wm["occupied_pct"],
            "leased_pct": wm["leased_pct"],
            "trend_30_day": wm["trend_30_day"],
            "trend_60_day": wm["trend_60_day"],
            "concessions": xlsx_data.get("concessions", []),
            "leased_num": wm["leased_num"],
            "occupied_num": wm["occupied_num"],
            "new_prospects": wm["new_prospects"],
            "walk_in_traffic": wm["walk_in_traffic"],
            "gross_leases": wm["gross_leases"],
            "net_leases": wm["net_leases"],
            "move_in": wm["move_in"],
            "move_out": wm["move_out"],
            "psf_all_leases": wm.get("psf_all_leases"),
            "psf_executed_effective": wm.get("psf_executed_effective"),
            # No supplemental DOCX data for Ancora
            "delinquency_total": None,
            "evictions_filed": None,
            "renewal_conversion_pct": None,
            "renewal_expiring_count": None,
            "trend_15_day": None,
            "notes": [],
        }
        weeks.append(week_entry)

    print(f"  [leasing] Total weeks in output: {len(weeks)}")

    result = {"weeks": weeks}
    if xlsx_data:
        result["xlsx_data"] = {
            "concessions": xlsx_data["concessions"],
            "construction_notes": xlsx_data.get("construction_notes", []),
            "delinquency_notes": xlsx_data["delinquency_notes"],
            # No expiration matrix for lease-up property
            "expiration_matrix": [],
        }

    return result
