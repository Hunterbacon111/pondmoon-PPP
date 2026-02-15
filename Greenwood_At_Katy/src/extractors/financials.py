"""Extract T-12 financial actuals from Greenwood 12 Month Statements.

Supports dual-year parsing: current year + prior year T-12 files for
year-over-year comparison.  The output JSON is backward-compatible —
top-level keys (status, source, period, months, metrics, annual_totals)
always point to the *current* year.  New keys `prior` and
`yoy_comparison` are appended when a second file is found.
"""
from src.config import DATA_T12, PROPERTY
import os
import glob
import re
from openpyxl import load_workbook


# Row-to-key mapping: (account_code_prefix or exact) -> key name
# We map subtotal/total rows by their account codes
ROW_MAP = {
    "41000-000": "market_rent",
    "41010-000": "gain_loss_to_lease",
    "41029-099": "potential_rent",
    "41091-000": "one_time_concessions",
    "41092-000": "renewal_concessions",
    "41093-000": "recurring_concessions",
    "41100-000": "vacancy_loss",
    "41110-000": "employee_units",
    "41115-000": "courtesy_officer_units",
    "41120-000": "model_storage_units",
    "41150-000": "bad_debt_rent",
    "41155-000": "bad_debt_recovery",
    "41999-098": "total_other_rental_inc",
    "41999-099": "total_rental_income",
    "43599-099": "total_other_income",
    "43999-099": "total_residential_income",
    "49999-999": "total_income",
    # Expenses
    "51599-099": "payroll_benefits",
    "52299-099": "repairs_maintenance",
    "52799-099": "make_ready",
    "52999-099": "recreational_amenities",
    "53298-099": "contract_services",
    "53999-099": "total_general_maintenance",
    "54999-099": "marketing",
    "58199-099": "office_expenses",
    "58398-099": "other_admin",
    "58399-099": "total_ga",
    "59999-099": "utilities",
    "61999-099": "management_fees",
    "62999-099": "taxes",
    "63999-099": "insurance",
    "66999-099": "total_non_recoverable_opex",
    "66999-199": "total_opex",
    "69999-090": "total_operating_expenses",
    "69999-099": "noi",
}

# Individual income items we also want for detail
DETAIL_INCOME_KEYS = {
    "43010-000": "admin_fees_income",
    "43020-000": "application_fees",
    "43080-000": "damages_income",
    "43105-000": "garage_income",
    "43125-000": "interest_income",
    "43135-000": "late_charge_fees",
    "43185-000": "package_locker_income",
    "43190-000": "parking_carport_income",
    "43260-000": "utility_reimbursement",
    "43261-000": "pest_control_rebill",
    "43262-000": "trash_rebill",
    "43263-000": "trash_door_to_door_rebill",
    "43264-001": "water_sewer_rebill",
}

# Keys used for YoY comparison (subset of all metrics)
YOY_KEYS = [
    "market_rent", "gain_loss_to_lease", "potential_rent",
    "one_time_concessions", "renewal_concessions", "vacancy_loss",
    "bad_debt_rent", "total_rental_income", "total_other_income",
    "total_income",
    "payroll_benefits", "repairs_maintenance", "make_ready",
    "contract_services", "marketing", "office_expenses", "other_admin",
    "utilities", "management_fees", "taxes", "insurance",
    "total_opex", "noi",
]


# ---------------------------------------------------------------------------
#  Helper: parse a single T-12 xlsx file
# ---------------------------------------------------------------------------
def _parse_t12_file(filepath):
    """Parse one T-12 xlsx and return structured dict."""
    fname = os.path.basename(filepath)
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        all_rows.append(list(row))
    wb.close()

    # Parse period string (row 3)
    period_str = str(all_rows[2][1] if all_rows[2][1] else all_rows[2][0] or "")
    if not period_str or "Period" not in period_str:
        for r in all_rows[:5]:
            for c in r:
                if c and "Period" in str(c):
                    period_str = str(c)
                    break

    # Month labels from header row (row 5, index 4)
    header_row = all_rows[4] if len(all_rows) > 4 else []
    months = []
    for ci in range(2, 14):
        val = header_row[ci] if ci < len(header_row) else None
        if val:
            if hasattr(val, 'strftime'):
                months.append(val.strftime('%b %Y'))
            else:
                months.append(str(val))
        else:
            months.append(f"M{ci-1}")

    # Parse data rows
    all_map = {**ROW_MAP, **DETAIL_INCOME_KEYS}
    metrics = {}

    for row in all_rows[5:]:
        if not row or len(row) < 3:
            continue
        acct_code = str(row[0]).strip() if row[0] else ""
        label = str(row[1]).strip() if row[1] else ""

        if acct_code in all_map:
            key = all_map[acct_code]
            monthly_vals = []
            for ci in range(2, 14):
                v = row[ci] if ci < len(row) else None
                if v is not None and isinstance(v, (int, float)):
                    monthly_vals.append(round(v, 2))
                else:
                    monthly_vals.append(0)

            annual_total = row[14] if len(row) > 14 and isinstance(row[14], (int, float)) else sum(monthly_vals)

            metrics[key] = monthly_vals
            metrics[f"{key}_label"] = label
            metrics[f"{key}_annual"] = round(annual_total, 2)

    # Derived: total concessions
    conc_1 = metrics.get("one_time_concessions", [0]*12)
    conc_2 = metrics.get("renewal_concessions", [0]*12)
    conc_3 = metrics.get("recurring_concessions", [0]*12)
    total_concessions = [round(conc_1[i] + conc_2[i] + conc_3[i], 2) for i in range(12)]
    metrics["total_concessions"] = total_concessions

    # Derived: vacancy + concessions
    vacancy = metrics.get("vacancy_loss", [0]*12)
    metrics["vacancy_and_concessions"] = [round(vacancy[i] + total_concessions[i], 2) for i in range(12)]

    total_units = PROPERTY["total_units"]
    annual_noi = sum(metrics.get("noi", [0]*12))
    annual_income = sum(metrics.get("total_income", [0]*12))
    annual_opex = sum(metrics.get("total_opex", [0]*12))

    annual_totals = {
        "total_income": round(annual_income, 0),
        "total_rental_income": round(sum(metrics.get("total_rental_income", [0]*12)), 0),
        "total_other_income": round(sum(metrics.get("total_other_income", [0]*12)), 0),
        "total_opex": round(annual_opex, 0),
        "noi": round(annual_noi, 0),
        "noi_per_unit": round(annual_noi / total_units, 0) if total_units else 0,
        "opex_ratio": round(annual_opex / annual_income * 100, 1) if annual_income else 0,
        "potential_rent": round(sum(metrics.get("potential_rent", [0]*12)), 0),
        "market_rent": round(sum(metrics.get("market_rent", [0]*12)), 0),
        "total_concessions": round(sum(total_concessions), 0),
        "vacancy_loss": round(sum(vacancy), 0),
        "bad_debt": round(sum(metrics.get("bad_debt_rent", [0]*12)), 0),
        "payroll_benefits": round(sum(metrics.get("payroll_benefits", [0]*12)), 0),
        "repairs_maintenance": round(sum(metrics.get("repairs_maintenance", [0]*12)), 0),
        "make_ready": round(sum(metrics.get("make_ready", [0]*12)), 0),
        "contract_services": round(sum(metrics.get("contract_services", [0]*12)), 0),
        "marketing": round(sum(metrics.get("marketing", [0]*12)), 0),
        "utilities": round(sum(metrics.get("utilities", [0]*12)), 0),
        "management_fees": round(sum(metrics.get("management_fees", [0]*12)), 0),
        "insurance": round(sum(metrics.get("insurance", [0]*12)), 0),
        "taxes": round(sum(metrics.get("taxes", [0]*12)), 0),
    }

    return {
        "source": fname,
        "period": period_str,
        "months": months,
        "metrics": metrics,
        "annual_totals": annual_totals,
        "_annual_noi": annual_noi,
        "_annual_income": annual_income,
        "_annual_opex": annual_opex,
    }


# ---------------------------------------------------------------------------
#  Helper: extract end-year from period string to determine current vs prior
# ---------------------------------------------------------------------------
def _period_end_year(period_str):
    """Extract the end year from a period string like 'Period = Oct 2024-Sep 2025'."""
    m = re.findall(r'(\d{4})', str(period_str))
    if m:
        return max(int(y) for y in m)
    return 0


# ---------------------------------------------------------------------------
#  Helper: compute year-over-year comparison
# ---------------------------------------------------------------------------
def _compute_yoy(current, prior):
    """Compute YoY comparison between current and prior T-12 data."""
    c_metrics = current["metrics"]
    p_metrics = prior["metrics"]
    c_totals = current["annual_totals"]
    p_totals = prior["annual_totals"]

    line_items = {}
    for key in YOY_KEYS:
        c_vals = c_metrics.get(key, [0]*12)
        p_vals = p_metrics.get(key, [0]*12)
        c_annual = sum(c_vals)
        p_annual = sum(p_vals)
        change_abs = round(c_annual - p_annual, 0)
        change_pct = round((c_annual - p_annual) / abs(p_annual) * 100, 1) if p_annual != 0 else None
        label = c_metrics.get(f"{key}_label", key.replace("_", " ").title())

        # Monthly changes
        monthly_change_abs = [round(c_vals[i] - p_vals[i], 0) for i in range(min(len(c_vals), len(p_vals)))]
        monthly_change_pct = []
        for i in range(min(len(c_vals), len(p_vals))):
            if p_vals[i] != 0:
                monthly_change_pct.append(round((c_vals[i] - p_vals[i]) / abs(p_vals[i]) * 100, 1))
            else:
                monthly_change_pct.append(None)

        line_items[key] = {
            "label": label,
            "current_annual": round(c_annual, 0),
            "prior_annual": round(p_annual, 0),
            "change_abs": change_abs,
            "change_pct": change_pct,
            "monthly_change_abs": monthly_change_abs,
            "monthly_change_pct": monthly_change_pct,
        }

    # Summary: biggest OpEx increases/decreases
    opex_keys = [
        "payroll_benefits", "repairs_maintenance", "make_ready",
        "contract_services", "marketing", "utilities",
        "management_fees", "insurance", "taxes",
    ]
    opex_changes = []
    for k in opex_keys:
        if k in line_items and line_items[k]["change_pct"] is not None:
            opex_changes.append({
                "key": k,
                "label": line_items[k]["label"],
                "change_abs": line_items[k]["change_abs"],
                "change_pct": line_items[k]["change_pct"],
            })

    opex_changes_sorted = sorted(opex_changes, key=lambda x: x["change_pct"], reverse=True)

    def _safe_pct(key):
        li = line_items.get(key, {})
        return li.get("change_pct")

    summary = {
        "total_income_change_abs": line_items.get("total_income", {}).get("change_abs", 0),
        "total_income_change_pct": _safe_pct("total_income"),
        "total_opex_change_abs": line_items.get("total_opex", {}).get("change_abs", 0),
        "total_opex_change_pct": _safe_pct("total_opex"),
        "noi_change_abs": line_items.get("noi", {}).get("change_abs", 0),
        "noi_change_pct": _safe_pct("noi"),
        "biggest_opex_increase": opex_changes_sorted[0] if opex_changes_sorted else None,
        "biggest_opex_decrease": opex_changes_sorted[-1] if opex_changes_sorted else None,
    }

    return {
        "line_items": line_items,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
#  Main entry point
# ---------------------------------------------------------------------------
def extract():
    """Extract T-12 financial actuals from xlsx file(s).

    Looks for all xlsx files in DATA_T12 directory.  Parses each, identifies
    current vs prior by end-year in period string.  Returns backward-compatible
    structure with optional `prior` and `yoy_comparison` keys.
    """
    if not os.path.exists(DATA_T12):
        print("  [financials] No T-12 directory found.")
        return {"status": "awaiting_data", "months": [], "metrics": {}}

    matches = glob.glob(os.path.join(DATA_T12, "*.xlsx"))
    matches = [f for f in matches if not os.path.basename(f).startswith("~")]

    if not matches:
        print("  [financials] No T-12 statement found.")
        return {"status": "awaiting_data", "months": [], "metrics": {}}

    # Parse all files
    parsed = []
    for fp in matches:
        try:
            data = _parse_t12_file(fp)
            data["_end_year"] = _period_end_year(data["period"])
            parsed.append(data)
            print(f"  [financials] Reading T-12: {data['source']}")
        except Exception as e:
            print(f"  [financials] Error reading {os.path.basename(fp)}: {e}")

    if not parsed:
        return {"status": "awaiting_data", "months": [], "metrics": {}}

    # Sort by end year descending: first = most recent = current
    parsed.sort(key=lambda x: x["_end_year"], reverse=True)
    current = parsed[0]
    prior = parsed[1] if len(parsed) > 1 else None

    print(f"  [financials] Period: {current['months'][0]} to {current['months'][-1]}")
    print(f"  [financials] T-12 NOI: ${current['_annual_noi']:,.0f} | "
          f"Income: ${current['_annual_income']:,.0f} | OpEx: ${current['_annual_opex']:,.0f}")

    # Build result (backward-compatible top-level keys)
    result = {
        "status": "loaded",
        "source": current["source"],
        "period": current["period"],
        "months": current["months"],
        "metrics": current["metrics"],
        "annual_totals": current["annual_totals"],
    }

    if prior:
        print(f"  [financials] Prior T-12: {prior['source']} "
              f"(NOI: ${prior['_annual_noi']:,.0f} | Income: ${prior['_annual_income']:,.0f})")
        result["prior"] = {
            "source": prior["source"],
            "period": prior["period"],
            "months": prior["months"],
            "metrics": prior["metrics"],
            "annual_totals": prior["annual_totals"],
        }
        result["yoy_comparison"] = _compute_yoy(current, prior)
    else:
        result["prior"] = None
        result["yoy_comparison"] = None

    return result
