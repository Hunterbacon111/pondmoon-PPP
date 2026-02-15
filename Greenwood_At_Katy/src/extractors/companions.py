"""Extract T-12 financial data from companion properties (non-Greenwood).

Each companion property may use a different PM system with a different
Excel format.  This module dispatches to format-specific parsers and
normalises the output to the same metric key names used by the Greenwood
T-12 extractor (financials.py) so the AI prompt can compare across
properties uniformly.
"""
import os
import re
import glob
from openpyxl import load_workbook
from src.config import COMPANION_PROPERTIES
from src.extractors.financials import _compute_yoy, YOY_KEYS


# ============================================================================
#  Trails at City Park — format-specific parser
# ============================================================================

# Account-code-based mapping:  Trails code -> standard key
TRAILS_ACCT_MAP = {
    "5120": "market_rent",
    "5210": "gain_loss_to_lease",
    "5220": "vacancy_loss",
    "5295": "total_concessions",
    "5297": "bad_debt_rent",
    "6720": "insurance",
}

# Label-based mapping for subtotals/totals (matched case-insensitively)
TRAILS_LABEL_MAP = {
    "total gross scheduled rent": "potential_rent",
    "total rental income": "total_rental_income",
    "total other revenue": "total_other_income",
    "total revenue": "total_income",
    "total payroll expenses": "payroll_benefits",
    "total management fee expenses": "management_fees",
    "total management fee": "management_fees",
    "total advertising expenses": "marketing",
    "total advertising": "marketing",
    "total administrative expenses": "office_expenses",
    "total administrative": "office_expenses",
    "total utilities expense": "utilities",
    "total utilities expenses": "utilities",
    "total turnover cost": "make_ready",
    "total turnover": "make_ready",
    "total contract services": "contract_services",
    "total operating & maintenance expenses": "repairs_maintenance",
    "total operating expenses": "total_opex",
    "net operating income": "noi",
    "net operating income (noi)": "noi",
}

# Taxes need special handling: sum 6710 + 6711
TRAILS_TAX_CODES = {"6710", "6711"}


def _parse_trails_t12(filepath, prop_info):
    """Parse a single Trails-format T-12 xlsx file.

    Returns dict with keys: source, period, months, metrics, annual_totals,
    _annual_noi, _annual_income, _annual_opex (internal).
    """
    fname = os.path.basename(filepath)
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        all_rows.append(list(row))
    wb.close()

    # --- Detect header end: find first row containing account "5120" ---
    data_start = 0
    for i, row in enumerate(all_rows):
        cell0 = str(row[0]).strip() if row and row[0] else ""
        if re.search(r'\b5120\b', cell0):
            data_start = i
            break

    # --- Parse month labels from header rows ---
    months = []
    # Look for a row above data_start that has date-like values in cols 1-12
    for i in range(max(0, data_start - 5), data_start):
        row = all_rows[i]
        if len(row) > 12:
            # Check if cols 1-12 look like dates or date strings
            sample = row[1] if len(row) > 1 else None
            if sample is not None:
                s = str(sample)
                if hasattr(sample, 'strftime'):
                    # datetime objects
                    months = []
                    for ci in range(1, 13):
                        v = row[ci] if ci < len(row) else None
                        if v and hasattr(v, 'strftime'):
                            months.append(v.strftime('%b %Y'))
                        else:
                            months.append(f"M{ci}")
                    break
                elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', s):
                    # String dates like "01/31/2025"
                    months = []
                    for ci in range(1, 13):
                        v = str(row[ci]) if ci < len(row) and row[ci] else ""
                        m = re.match(r'(\d{1,2})/\d{1,2}/(\d{4})', v)
                        if m:
                            month_num = int(m.group(1))
                            year = m.group(2)
                            month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                            months.append(f"{month_names[month_num-1]} {year}")
                        else:
                            months.append(f"M{ci}")
                    break

    if not months:
        months = [f"M{i+1}" for i in range(12)]

    # Detect if months are in reverse chronological order (Dec → Jan)
    # If so, we'll need to reverse both months and monthly data values
    reverse_columns = False
    if len(months) >= 2:
        # Parse year-month from first and last month labels
        m1 = re.search(r'(\w+)\s+(\d{4})', months[0])
        m2 = re.search(r'(\w+)\s+(\d{4})', months[-1])
        if m1 and m2:
            month_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            y1, y2 = int(m1.group(2)), int(m2.group(2))
            mi1 = month_order.index(m1.group(1)) if m1.group(1) in month_order else 0
            mi2 = month_order.index(m2.group(1)) if m2.group(1) in month_order else 0
            if y1 > y2 or (y1 == y2 and mi1 > mi2):
                reverse_columns = True
                months = list(reversed(months))

    # Build period string from months
    period_str = f"{months[0]} - {months[-1]}" if months else ""

    # --- Parse data rows ---
    metrics = {}
    tax_monthly = [0.0] * 12  # accumulate tax codes

    for row in all_rows[data_start:]:
        if not row or len(row) < 2:
            continue

        cell0 = str(row[0]).strip() if row[0] else ""
        if not cell0:
            continue

        # Extract monthly values from cols 1-12
        monthly_vals = []
        for ci in range(1, 13):
            v = row[ci] if ci < len(row) else None
            if v is not None and isinstance(v, (int, float)):
                monthly_vals.append(round(v, 2))
            else:
                monthly_vals.append(0)

        # If columns are in reverse chronological order, flip to chronological
        if reverse_columns:
            monthly_vals = list(reversed(monthly_vals))

        # Annual total from col 13
        annual_total = row[13] if len(row) > 13 and isinstance(row[13], (int, float)) else sum(monthly_vals)

        # Try account code match first
        acct_match = re.match(r'\s*(\d{4}(?:\.\d{3})?)\s*-\s*(.+)', cell0)
        matched_key = None

        if acct_match:
            code = acct_match.group(1)
            label = acct_match.group(2).strip()

            if code in TRAILS_ACCT_MAP:
                matched_key = TRAILS_ACCT_MAP[code]
                metrics[matched_key] = monthly_vals
                metrics[f"{matched_key}_label"] = label
                metrics[f"{matched_key}_annual"] = round(annual_total, 2)

            if code in TRAILS_TAX_CODES:
                # Accumulate taxes
                for i in range(12):
                    tax_monthly[i] = round(tax_monthly[i] + monthly_vals[i], 2)

        # Try label-based match (for subtotals like "Total Revenue")
        if matched_key is None:
            cell0_lower = cell0.lower().strip()
            for pattern, key in TRAILS_LABEL_MAP.items():
                if cell0_lower == pattern or cell0_lower.startswith(pattern):
                    if key not in metrics:  # avoid overwriting if already set
                        metrics[key] = monthly_vals
                        metrics[f"{key}_label"] = cell0.strip()
                        metrics[f"{key}_annual"] = round(annual_total, 2)
                    break

    # Store accumulated taxes
    metrics["taxes"] = tax_monthly
    metrics["taxes_label"] = "Real Estate & Franchise Tax"
    metrics["taxes_annual"] = round(sum(tax_monthly), 2)

    # --- Compute annual totals ---
    total_units = prop_info["total_units"]
    annual_noi = sum(metrics.get("noi", [0]*12))
    annual_income = sum(metrics.get("total_income", [0]*12))
    annual_opex = sum(metrics.get("total_opex", [0]*12))

    annual_totals = {
        "total_income": round(annual_income, 0),
        "total_rental_income": round(sum(metrics.get("total_rental_income", [0]*12)), 0),
        "total_other_income": round(sum(metrics.get("total_other_income", [0]*12)), 0),
        "total_opex": round(annual_opex, 0),
        "noi": round(annual_noi, 0),
        "noi_per_unit": round(annual_noi / total_units, 0) if total_units else 0,
        "opex_ratio": round(annual_opex / annual_income * 100, 1) if annual_income else 0,
        "potential_rent": round(sum(metrics.get("potential_rent", [0]*12)), 0),
        "market_rent": round(sum(metrics.get("market_rent", [0]*12)), 0),
        "total_concessions": round(sum(metrics.get("total_concessions", [0]*12)), 0),
        "vacancy_loss": round(sum(metrics.get("vacancy_loss", [0]*12)), 0),
        "bad_debt": round(sum(metrics.get("bad_debt_rent", [0]*12)), 0),
        "payroll_benefits": round(sum(metrics.get("payroll_benefits", [0]*12)), 0),
        "repairs_maintenance": round(sum(metrics.get("repairs_maintenance", [0]*12)), 0),
        "make_ready": round(sum(metrics.get("make_ready", [0]*12)), 0),
        "contract_services": round(sum(metrics.get("contract_services", [0]*12)), 0),
        "marketing": round(sum(metrics.get("marketing", [0]*12)), 0),
        "utilities": round(sum(metrics.get("utilities", [0]*12)), 0),
        "management_fees": round(sum(metrics.get("management_fees", [0]*12)), 0),
        "insurance": round(sum(metrics.get("insurance", [0]*12)), 0),
        "taxes": round(sum(metrics.get("taxes", [0]*12)), 0),
    }

    return {
        "source": fname,
        "period": period_str,
        "months": months,
        "metrics": metrics,
        "annual_totals": annual_totals,
        "_annual_noi": annual_noi,
        "_annual_income": annual_income,
        "_annual_opex": annual_opex,
    }


def _detect_year_from_filename(fname):
    """Extract year from filename like 'T-12 2025.12 Trails.xlsx' -> 2025."""
    m = re.search(r'(\d{4})', fname)
    return int(m.group(1)) if m else 0


# ============================================================================
#  Main entry point
# ============================================================================

def extract():
    """Extract T-12 data for all companion properties.

    Returns dict keyed by property slug:
    {
        "trails": {
            "name": ..., "short_name": ..., "total_units": ..., "total_sf": ...,
            "current": { source, period, months, metrics, annual_totals },
            "prior":   { ... } or null,
            "yoy_comparison": { ... } or null,
        }
    }
    """
    result = {}

    for slug, prop_info in COMPANION_PROPERTIES.items():
        data_dir = prop_info["data_dir"]
        if not os.path.exists(data_dir):
            print(f"  [companions] {prop_info['name']}: data directory not found, skipping.")
            continue

        # Find xlsx files
        matches = glob.glob(os.path.join(data_dir, "*.xlsx"))
        matches = [f for f in matches if not os.path.basename(f).startswith("~")]

        if not matches:
            print(f"  [companions] {prop_info['name']}: no T-12 files found, skipping.")
            continue

        # Select parser based on format
        fmt = prop_info.get("format", "trails")

        parsed = []
        for fp in matches:
            try:
                if fmt == "trails":
                    data = _parse_trails_t12(fp, prop_info)
                else:
                    print(f"  [companions] Unknown format '{fmt}' for {prop_info['name']}, skipping {os.path.basename(fp)}")
                    continue

                data["_year"] = _detect_year_from_filename(os.path.basename(fp))
                parsed.append(data)
                print(f"  [companions] {prop_info['short_name']}: parsed {data['source']}")
            except Exception as e:
                print(f"  [companions] Error parsing {os.path.basename(fp)}: {e}")

        if not parsed:
            continue

        # Sort by year descending: first = most recent = current
        parsed.sort(key=lambda x: x["_year"], reverse=True)
        current = parsed[0]
        prior = parsed[1] if len(parsed) > 1 else None

        print(f"  [companions] {prop_info['short_name']}: "
              f"NOI ${current['_annual_noi']:,.0f} | "
              f"Income ${current['_annual_income']:,.0f} | "
              f"OpEx ${current['_annual_opex']:,.0f}")

        prop_result = {
            "name": prop_info["name"],
            "short_name": prop_info["short_name"],
            "total_units": prop_info["total_units"],
            "total_sf": prop_info["total_sf"],
            "current": {
                "source": current["source"],
                "period": current["period"],
                "months": current["months"],
                "metrics": current["metrics"],
                "annual_totals": current["annual_totals"],
            },
        }

        if prior:
            print(f"  [companions] {prop_info['short_name']} prior: "
                  f"NOI ${prior['_annual_noi']:,.0f} | Income ${prior['_annual_income']:,.0f}")
            prop_result["prior"] = {
                "source": prior["source"],
                "period": prior["period"],
                "months": prior["months"],
                "metrics": prior["metrics"],
                "annual_totals": prior["annual_totals"],
            }
            prop_result["yoy_comparison"] = _compute_yoy(current, prior)
        else:
            prop_result["prior"] = None
            prop_result["yoy_comparison"] = None

        result[slug] = prop_result

    if not result:
        print("  [companions] No companion property data found.")

    return result
