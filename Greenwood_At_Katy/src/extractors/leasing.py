"""Extract weekly leasing data from XLSX (primary) and DOCX/PDF (supplemental)."""
import os
import re
import glob
from datetime import datetime
from docx import Document
from src.config import DATA_LEASING


def _parse_date_from_filename(filename):
    """Extract date from filename like 'GWK Weekly Update Week Ending 2.11.26.docx'."""
    match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', filename)
    if match:
        month, day, year = match.groups()
        year = int(year)
        if year < 100:
            year += 2000
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return None


def _extract_pct(text):
    """Extract a percentage value from text like '87.04%'."""
    match = re.search(r'(\d+\.?\d*)%', text)
    if match:
        return float(match.group(1))
    return None


def _extract_dollar(text):
    """Extract a dollar amount from text like '$191,140.95'."""
    match = re.search(r'\$?([\d,]+\.?\d*)', text)
    if match:
        return float(match.group(1).replace(',', ''))
    return None


def _extract_int(text):
    """Extract an integer from text like 'filed on 5 residents'."""
    match = re.search(r'(\d+)', text)
    if match:
        return int(match.group(1))
    return None


def _parse_xlsx_weekly(filepath):
    """Parse the Pondmoon weekly tracking XLSX file.

    Returns a dict with:
      - weekly_metrics: list of weekly snapshots (newest first in sheet, we reverse)
      - concessions: list of concession strings
      - expiration_matrix: list of per-date expiration data
      - delinquency_notes: list of notes
      - psf_all_leases: rent PSF across all leases
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Greenwood"]

    # --- Top table: rows 3-17, columns C onward ---
    # Row 3 = dates, Row 4 = Leased#, Row 5 = Leased%, etc.
    dates = []
    col_start = 3  # column C
    for col in range(col_start, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val and isinstance(val, datetime):
            dates.append((col, val))

    weekly_metrics = []
    for col, dt in dates:
        def cell_val(row):
            v = ws.cell(row=row, column=col).value
            return v

        entry = {
            "week_ending": dt.strftime("%Y-%m-%d"),
            "leased_num": cell_val(4),
            "leased_pct": cell_val(5),
            "occupied_num": cell_val(6),
            "occupied_pct": cell_val(7),
            "move_out": cell_val(8),
            "move_in": cell_val(9),
            "new_prospects": cell_val(10),
            "walk_in_traffic": cell_val(11),
            "gross_leases": cell_val(12),
            "net_leases": cell_val(13),
            "trend_30_day": cell_val(14),
            "trend_60_day": cell_val(15),
            "psf_all_leases": cell_val(16),
            "psf_this_week": cell_val(17),
        }

        # Convert percentages stored as decimals (e.g. 0.8981 → 89.81)
        for key in ["leased_pct", "occupied_pct", "trend_30_day", "trend_60_day"]:
            if entry[key] is not None and isinstance(entry[key], (int, float)):
                entry[key] = round(entry[key] * 100, 2)

        # Round unit counts
        for key in ["leased_num", "occupied_num"]:
            if entry[key] is not None and isinstance(entry[key], (int, float)):
                entry[key] = round(entry[key])

        weekly_metrics.append(entry)

    # Fix date entry errors (e.g., 2026-12-28 should be 2025-12-28)
    # If a date is out of sequence and changing the year fixes it, apply the fix
    raw_dates = [datetime.strptime(w["week_ending"], "%Y-%m-%d") for w in weekly_metrics]
    if len(raw_dates) > 2:
        # Sort by column position (original order in sheet is newest-first)
        # Check each date: if it's clearly out of sequence, try year-1
        sorted_by_val = sorted(range(len(raw_dates)), key=lambda i: raw_dates[i])
        for i, wm in enumerate(weekly_metrics):
            dt = datetime.strptime(wm["week_ending"], "%Y-%m-%d")
            # If this date is more than 6 months after the max of all other dates, it's likely a year error
            others = [datetime.strptime(w["week_ending"], "%Y-%m-%d") for j, w in enumerate(weekly_metrics) if j != i]
            if others:
                max_other = max(others)
                min_other = min(others)
                if dt > max_other and (dt - max_other).days > 180:
                    fixed = dt.replace(year=dt.year - 1)
                    if min_other <= fixed <= max_other:
                        wm["week_ending"] = fixed.strftime("%Y-%m-%d")
                        print(f"  [leasing] Fixed date: {dt.strftime('%Y-%m-%d')} -> {wm['week_ending']}")

    # Sort chronologically (oldest first)
    weekly_metrics.sort(key=lambda w: w["week_ending"])

    # --- Concessions (rows 20-22) ---
    concessions = []
    for row in range(20, 23):
        val = ws.cell(row=row, column=2).value  # column B
        if val and isinstance(val, str) and val.strip() and not val.strip().lower().startswith("concession"):
            # Skip locator commission — not a concession
            if "locator commission" in val.strip().lower():
                continue
            concessions.append(val.strip())

    # --- Expiration Matrix (rows 25-34) ---
    # Each "date block" has 3 month columns (e.g., JAN, FEB, MAR)
    # Row 25 has the dates, Row 26 has month labels
    # Rows 27-34 have: Expiring, Notice, Renewals, MTM, Transfers, Pending, Avg $ Increase %, Renewals Retention
    exp_dates = []
    exp_col = 3  # start from column C
    while exp_col <= ws.max_column:
        date_val = ws.cell(row=25, column=exp_col).value
        if date_val is None:
            # Check next columns
            exp_col += 1
            if exp_col > ws.max_column:
                break
            continue
        if isinstance(date_val, datetime):
            exp_dates.append((exp_col, date_val))
        exp_col += 1

    # Read the unique date blocks from row 25
    # Each date in row 25 is in the MIDDLE of a 3-column block
    # e.g., date in Col 4 → block is Col 3 (month1), Col 4 (month2), Col 5 (month3)
    expiration_matrix = []
    date_positions = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=25, column=col).value
        if val and isinstance(val, datetime):
            date_positions.append((col, val))

    for date_col, dt in date_positions:
        block = {
            "as_of_date": dt.strftime("%Y-%m-%d"),
            "months": []
        }
        # Block: col-1, col, col+1
        for c in [date_col - 1, date_col, date_col + 1]:
            if c < 1 or c > ws.max_column:
                continue
            month_label = ws.cell(row=26, column=c).value
            if month_label is None:
                continue

            def get_exp_val(row, col=c):
                v = ws.cell(row=row, column=col).value
                return v

            retention_val = get_exp_val(34)
            if isinstance(retention_val, (int, float)):
                retention_val = round(retention_val * 100, 1)

            avg_increase_val = get_exp_val(33)
            if isinstance(avg_increase_val, (int, float)) and avg_increase_val < 1:
                avg_increase_val = round(avg_increase_val * 100, 2)

            month_data = {
                "month": month_label,
                "expiring": get_exp_val(27),
                "notice": get_exp_val(28),
                "renewals": get_exp_val(29),
                "month_to_month": get_exp_val(30),
                "transfers": get_exp_val(31),
                "pending": get_exp_val(32),
                "avg_increase_pct": avg_increase_val if avg_increase_val != '-' else None,
                "retention_pct": retention_val if retention_val != '-' else None,
            }
            block["months"].append(month_data)
        expiration_matrix.append(block)

    # --- Additional Notes (rows 36-39) ---
    # Row 36 is typically a section header ("Delinquency"), rows 37+ are additional notes
    delinquency_notes = []
    for row in range(36, 40):
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, str) and val.strip():
            label = val.strip()
            # Skip bare section headers
            if label.lower() in ("delinquency", "delinquency:", "concessions:", "concessions"):
                continue
            delinquency_notes.append(label)

    return {
        "weekly_metrics": weekly_metrics,
        "concessions": concessions,
        "expiration_matrix": expiration_matrix,
        "delinquency_notes": delinquency_notes,
    }


def _parse_docx(filepath):
    """Parse a weekly update DOCX file (supplemental details)."""
    doc = Document(filepath)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

    result = {
        "source_type": "docx",
        "occupancy_pct": None,
        "leased_pct": None,
        "trend_15_day": None,
        "trend_30_day": None,
        "trend_60_day": None,
        "concessions": [],
        "renewal_conversion_pct": None,
        "renewal_expiring_count": None,
        "delinquency_total": None,
        "evictions_filed": None,
        "notes": [],
    }

    section = None
    concession_collecting = False

    for i, para in enumerate(paragraphs):
        lower = para.lower()

        # Detect sections
        if lower.startswith("occupancy"):
            section = "occupancy"
            concession_collecting = False
        elif "current concession" in lower or "concession" in lower.split(":")[0].lower():
            section = "concessions"
            concession_collecting = True
            continue
        elif lower.startswith("renewal"):
            section = "renewal"
            concession_collecting = False
        elif lower.startswith("delinquency"):
            section = "delinquency"
            concession_collecting = False
        elif lower.startswith("follow up") or lower.startswith("follow-up"):
            section = "followup"
            concession_collecting = False
        elif lower.startswith("marketing") or lower.startswith("market update"):
            section = "marketing"
            concession_collecting = False

        # Parse based on current section
        if section == "occupancy":
            if "current:" in lower or "current :" in lower:
                result["occupancy_pct"] = _extract_pct(para)
            elif "leased:" in lower or "leased :" in lower:
                result["leased_pct"] = _extract_pct(para)
            elif "15 day" in lower or "15-day" in lower:
                result["trend_15_day"] = _extract_pct(para)
            elif "30 day" in lower or "30-day" in lower:
                result["trend_30_day"] = _extract_pct(para)
            elif "60 day" in lower or "60-day" in lower:
                result["trend_60_day"] = _extract_pct(para)

        elif section == "concessions" and concession_collecting:
            if para and not any(para.lower().startswith(s) for s in
                               ["renewal", "delinquency", "follow", "marketing", "market"]):
                result["concessions"].append(para)
            else:
                concession_collecting = False

        elif section == "renewal":
            if "conversion" in lower:
                pct = _extract_pct(para)
                if pct:
                    result["renewal_conversion_pct"] = pct
            if "expiring" in lower:
                result["renewal_expiring_count"] = _extract_int(para)

        elif section == "delinquency":
            if "$" in para and result["delinquency_total"] is None:
                amt = _extract_dollar(para)
                if amt and amt > 1000:
                    result["delinquency_total"] = amt
            if "eviction" in lower:
                result["evictions_filed"] = _extract_int(para)

        elif section == "followup":
            sig_keywords = ["best,", "community manager", "apartment homes", "greystar",
                            "please consider", "redefining", "@greystar", "832-"]
            if not any(kw in lower for kw in sig_keywords):
                result["notes"].append(para)

    return result


def extract():
    """Scan Data_Leasing directory and extract all weekly data.

    Primary source: Pondmoon weekly XLSX tracking file.
    Supplemental: individual DOCX/PDF weekly updates (for delinquency, notes).
    """
    if not os.path.exists(DATA_LEASING):
        print("  [leasing] Data_Leasing directory not found.")
        return {"weeks": [], "xlsx_data": None}

    # --- 1. Primary: XLSX weekly tracking file ---
    xlsx_files = glob.glob(os.path.join(DATA_LEASING, "**", "*Pondmoon*.xlsx"), recursive=True)
    if not xlsx_files:
        xlsx_files = glob.glob(os.path.join(DATA_LEASING, "**", "*Greenwood Weekly*.xlsx"), recursive=True)

    xlsx_data = None
    if xlsx_files:
        xlsx_path = xlsx_files[0]  # Use most recent
        print(f"  [leasing] Parsing PRIMARY XLSX: {os.path.basename(xlsx_path)}")
        xlsx_data = _parse_xlsx_weekly(xlsx_path)
        print(f"  [leasing] Found {len(xlsx_data['weekly_metrics'])} weekly snapshots")
        print(f"  [leasing] Found {len(xlsx_data['expiration_matrix'])} expiration matrix snapshots")
    else:
        print("  [leasing] No Pondmoon XLSX file found.")

    # --- 2. Supplemental: DOCX weekly updates ---
    docx_files = glob.glob(os.path.join(DATA_LEASING, "**", "*.docx"), recursive=True)
    # Exclude the xlsx-related files
    docx_files = [f for f in docx_files if "Pondmoon" not in os.path.basename(f)
                  and "Greenwood Weekly_Managed" not in os.path.basename(f)]

    supplemental = []
    for filepath in sorted(docx_files):
        filename = os.path.basename(filepath)
        date_str = _parse_date_from_filename(filename)
        if not date_str:
            print(f"  [leasing] Could not parse date from: {filename}")
            continue
        print(f"  [leasing] Parsing supplemental DOCX: {filename} -> {date_str}")
        data = _parse_docx(filepath)
        data["week_ending"] = date_str
        data["source_file"] = filename
        supplemental.append(data)

    supplemental.sort(key=lambda w: w["week_ending"])

    # --- 3. Build unified output ---
    # Convert xlsx weekly_metrics to the legacy "weeks" format for backward compatibility
    # but also include the full xlsx_data for the new dashboard tables
    weeks = []
    if xlsx_data:
        for wm in xlsx_data["weekly_metrics"]:
            week_entry = {
                "week_ending": wm["week_ending"],
                "source_type": "xlsx",
                "occupancy_pct": wm["occupied_pct"],
                "leased_pct": wm["leased_pct"],
                "trend_30_day": wm["trend_30_day"],
                "trend_60_day": wm["trend_60_day"],
                "concessions": xlsx_data.get("concessions", []),
                "leased_num": wm["leased_num"],
                "occupied_num": wm["occupied_num"],
                "new_prospects": wm["new_prospects"],
                "walk_in_traffic": wm["walk_in_traffic"],
                "gross_leases": wm["gross_leases"],
                "net_leases": wm["net_leases"],
                "move_in": wm["move_in"],
                "move_out": wm["move_out"],
                "psf_all_leases": wm.get("psf_all_leases"),
            }

            # Merge supplemental DOCX data if we have a matching date
            matching_docx = next((s for s in supplemental if s["week_ending"] == wm["week_ending"]), None)
            if matching_docx:
                week_entry["delinquency_total"] = matching_docx.get("delinquency_total")
                week_entry["evictions_filed"] = matching_docx.get("evictions_filed")
                week_entry["renewal_conversion_pct"] = matching_docx.get("renewal_conversion_pct")
                week_entry["renewal_expiring_count"] = matching_docx.get("renewal_expiring_count")
                week_entry["trend_15_day"] = matching_docx.get("trend_15_day")
                week_entry["notes"] = matching_docx.get("notes", [])
            else:
                week_entry["delinquency_total"] = None
                week_entry["evictions_filed"] = None
                week_entry["renewal_conversion_pct"] = None
                week_entry["renewal_expiring_count"] = None
                week_entry["trend_15_day"] = None
                week_entry["notes"] = []

            weeks.append(week_entry)

    # Add any DOCX-only weeks that aren't in xlsx
    xlsx_dates = set(w["week_ending"] for w in weeks)
    for s in supplemental:
        if s["week_ending"] not in xlsx_dates:
            weeks.append(s)

    weeks.sort(key=lambda w: w["week_ending"])
    print(f"  [leasing] Total weeks in output: {len(weeks)}")

    result = {"weeks": weeks}
    if xlsx_data:
        result["xlsx_data"] = {
            "concessions": xlsx_data["concessions"],
            "expiration_matrix": xlsx_data["expiration_matrix"],
            "delinquency_notes": xlsx_data["delinquency_notes"],
        }

    return result
